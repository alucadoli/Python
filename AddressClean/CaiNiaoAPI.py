#!/usr/bin/env python
# _*_ coding:utf-8 _*_
import sys
import getopt
import requests
import hashlib
import base64
import json
import os
from openpyxl import load_workbook
from openpyxl import Workbook

# url = "https://shidclink.cainiao.com/gateway/link.do"
# keys = "jfiaodkfIUHFHIE/.83("

def get_param(sign,content):
    payload = {
        'logistics_interface': content,
        'msg_type': 'ADDRESS_COMMON_PARSE',
        'logistic_provider_id': 'ADDRESS_ONLINE',
        'data_digest': sign
    }
    return payload

def get_data_digest(inputs,keys):
    m1 = hashlib.md5()
    m1.update((inputs+keys).encode('utf-8'))

    return base64.b64encode(m1.digest())

def set_reqstr(addr):
    stradd="<request><addressList><str>"+addr.strip()+"</str></addressList></request>"
    return stradd

def writetoexcel(path, result):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Addresses'
    sheet.append(list(result.keys()))
    sheet.append(list(result.values()))
    wb.save(path)
    print('Success to write!')

def readexcel(path,sheetname):
    wb = load_workbook(path)
    ws = wb[sheetname]
    list1 = []
    templist = []
    for row in ws.rows:
        for cell in row:
            templist.append(cell.value)
        list1.append(templist.copy())
        templist.clear()
    wb.close()
    return list1

def query_postcode(payload):

    url = "http://cpdc.chinapost.com.cn/web/index.php"

    postcode_headers = {
        'Accept': 'application/json, text/javascript, */*; q=0.01',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
        'Cookie': 'Hm_lvt_f9156f80cc764fbe69ecafe0e5202e63=1523152264; '
                  'Hm_lpvt_f9156f80cc764fbe69ecafe0e5202e63=1523152264; '
                  'UM_distinctid=162a2f31d1e2e5-0c9cf6e54c9d92-b34356b-144000-162a2f31d2649f; '
                  'PHPSESSID=bktioe6kcjne6059131aj8av73; '
                  'iptac-5C838FE1DDB7-FCZ1949N005=434f5250464353494e545c636e616c6c69404643535f4144--1523155850--'
                  '2507--1522814673--fb4d5c647d65af49333313cf52a6b668f4210057a2990c81b81c605ec69e425c',
        'DNT': '1',
        'Host': 'cpdc.chinapost.com.cn',
        'Proxy-Connection': 'keep-alive',
        'Referer': 'http://cpdc.chinapost.com.cn/web/',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                      'Chrome/65.0.3325.181 Safari/537.36',
        'X-Requested-With': 'XMLHttpRequest'
    }

    postcode_payload = {'m': 'postsearch', 'c': 'index', 'a': 'ajax_addr', 'searchkey': payload['searchkey'],
                        'flag': '1523156224735', 'provinceid': payload['provinceid'],
                        'cityid': payload['cityid'] + ' - r', 'areaid': payload['areaid'], 'dosubmit': '1'}
    try:
        result = requests.get(url, headers=postcode_headers, params=postcode_payload, verify=False)
        result.encoding = 'utf-8'
        return  result.json()['rs'][0]['POSTCODE']
    except Exception as e:
        return '000000'
        # if(result.json()['rs']):
        #     return result.json()['rs'][0]['POSTCODE']
        # else:
        #     return '0'

def call_address_split(addr,url="https://shidclink.cainiao.com/gateway/link.do",keys="jfiaodkfIUHFHIE/.83("):
    proxies = {
        "https": "http://10.3.246.5:8500"
    }

    headers = {
        'Accept': 'application / json, text / javascript, * / *; q = 0.01',
        'Accept - Encoding': 'gzip, deflate, br',
        'Accept - Language': 'zh - CN, zh;q = 0.9, en;q = 0.8',
        'Connection': 'keep - alive',
        'Content - Length': '258',
        'Content - Type': 'application / x - www - form - urlencoded;charset = UTF - 8',
        'DNT': '1',
        'Host': 'shidclink.cainiao.com',
        'Origin': 'https: // cloud.cainiao.com',
        'Referer': 'https: // cloud.cainiao.com / markets / cnwww / cncloud - dzk - detail - correct',
        'User - Agent': 'Mozilla / 5.0(Windows NT 10.0; Win64;x64) AppleWebKit / 537.36(KHTML, like Gecko) '
                        'Chrome / 64.0.3282.186 Safari / 537.36'
    }

    input_addr = set_reqstr(addr)
    result = json.loads(requests.post(url, headers=headers, data=get_param(get_data_digest(input_addr, keys), input_addr)).text)
    return result

#adict:result / stdict: template dict
def key_map(adict,stdict):
    for k,v in adict.items():
        stdict[k]=adict[k]
    return stdict

##########################################################################################################################
def main(argv):
    spath = ''
    sheetname = ''
    dpath = ''
    opts, args = getopt.getopt(sys.argv[1:], "hs:t:d:", ["help","spath=","sheetname=","dpath="])
    for opt, arg in opts:
        if opt == '-h':
            print('CaiNiaoAPI.py -s <inputfile> -t <excelsheetname> -d <outputfile>')
            sys.exit()
        elif opt in ("-s","--spath"):
            spath = arg
        elif opt in ("-t","--sheetname"):
            sheetname = arg
        elif opt in ("-d","--outputfile"):
            dpath = arg

    # path = r'D:\PythonProjects\AddressClean\iScalaAddr.xlsx'
    tmp_dict = {'code': '','saddress': '','mergeaddress':'','address': '','lng': '', 'town': '', 'city': '', 'cityId': '', 'townId': '', 'detailAddr': '',
                'districtId': '', 'district': '', 'id': '', 'prov': '', 'provId': '', 'lat': '', 'postcode': '','result':''}

    # Loop addresses to query
    # Generate Result xlsx file.
    wbresult = Workbook()
    ws1 = wbresult.active
    ws1.title = 'Standard Addresses Split Result'
    ws1.append(list(tmp_dict.keys()))
    # wbresult.save(r'D:\PythonProjects\AddressClean\SplitAddr.xlsx')
    wbresult.save(dpath)
    source = readexcel(spath,sheetname)[1:]
    totalrows = len(source)
    i=0
    for row in source:
        if(len(row)==0):break
        i += 1
        print('current record is %d, %10.2f%% in progress' %(i,i/totalrows*100))
        addr = set_reqstr(row[1])
        try:
            result = call_address_split(addr)
            if(result['data']):
                rs = key_map(result['data'][0],tmp_dict)
                rs['code'] = row[0]
                rs['saddress'] = row[1]
                rs['mergeaddress'] = rs['prov']+rs['city']+rs['district']+rs['town']+rs['detailAddr']
                payload = {'searchkey': row[1], 'provinceid': rs['provId'],
                           'cityid': rs['cityId'],'areaid': rs['districtId']}
                postcode = query_postcode(payload)
                rs['postcode'] = postcode
                rs['result'] = result['status'][0]
                ws1.append(list(rs.values()))
                if (i % 2 == 0):
                    wbresult.save(dpath)
        except Exception as e:
            print('Error Msg is :',e)
            print(row, result)
            ws1.append(row + list(str(result).split(' ')))
            wbresult.save(dpath)
    wbresult.save(dpath)

if __name__ == "__main__":
    main(sys.argv[1])






