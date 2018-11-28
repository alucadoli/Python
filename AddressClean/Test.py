#!/usr/bin/env python
# _*_ coding:utf-8 _*_
from openpyxl import load_workbook
from openpyxl import Workbook


path1=r'D:\PythonProjects\AddressClean\t.xlsx'

def test(path):
    wb = load_workbook(path)
    sheet = wb['Sheet1']
    sheet.cell(row=1,column=1,value='a')
    wb.save()


wbresult = Workbook()
ws1 = wbresult.active
ws1.title = 'Sheet1'
wbresult.save(r'D:\PythonProjects\AddressClean\t.xlsx')


wb = load_workbook(path1)
ws = wb['Sheet1']
for i in range(10):
    ws.cell(row=i+1,column=1,value=i+1)
    wb.save(path1)


